import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# SHEET 1: LOGFRAME BUDGET
# ============================================================
ws = wb.active
ws.title = "Logframe Budget"

dark_green_fill = PatternFill(start_color="2D5016", end_color="2D5016", fill_type="solid")
earth_fill = PatternFill(start_color="1A1A0E", end_color="1A1A0E", fill_type="solid")
light_green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
subtotal_fill = PatternFill(start_color="F0F4E8", end_color="F0F4E8", fill_type="solid")
year_total_fill = PatternFill(start_color="E8F0DD", end_color="E8F0DD", fill_type="solid")
grand_total_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
grand_row_fill = PatternFill(start_color="1A1A0E", end_color="1A1A0E", fill_type="solid")

white_font = Font(color="FFFFFF", bold=True, size=10)
white_font_sm = Font(color="FFFFFF", bold=True, size=9)
header_font = Font(bold=True, size=10)
sub_font = Font(size=9, color="4E4E3A")
money_font = Font(name="Consolas", size=9)
money_bold = Font(name="Consolas", size=9, bold=True, color="2D5016")

thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_al = Alignment(horizontal='right', vertical='center')
usd_fmt = '$#,##0'

# Title rows
ws.merge_cells('A1:R1')
ws['A1'] = "MooMe IoT-AI Livestock Monitoring Platform"
ws['A1'].font = Font(bold=True, size=14, color="2D5016"); ws['A1'].alignment = center

ws.merge_cells('A2:R2')
ws['A2'] = "Logframe: Quarterly Budget Breakdown | Technical Support / IoT (PI/E)"
ws['A2'].font = Font(size=11, color="4E4E3A"); ws['A2'].alignment = center

ws.merge_cells('A3:R3')
ws['A3'] = "Solution Design, Development & Promotion | Grand Total: $723,287 USD across 3 years"
ws['A3'].font = Font(size=10, color="7A7A62"); ws['A3'].alignment = center

# Headers row 5-6
ws.merge_cells('A5:A6'); ws['A5'] = 'Category'; ws['A5'].font = white_font; ws['A5'].fill = dark_green_fill; ws['A5'].alignment = center; ws['A5'].border = thin_border
ws.merge_cells('B5:B6'); ws['B5'] = 'Activity / Deliverable'; ws['B5'].font = white_font; ws['B5'].fill = dark_green_fill; ws['B5'].alignment = center; ws['B5'].border = thin_border

ws.merge_cells('C5:F5'); ws['C5'] = 'Year 1: $205,479'; ws['C5'].font = white_font; ws['C5'].fill = earth_fill; ws['C5'].alignment = center; ws['C5'].border = thin_border
ws.merge_cells('G5:G6'); ws['G5'] = 'Y1 Total'; ws['G5'].font = white_font; ws['G5'].fill = dark_green_fill; ws['G5'].alignment = center; ws['G5'].border = thin_border

ws.merge_cells('H5:K5'); ws['H5'] = 'Year 2: $246,575'; ws['H5'].font = white_font; ws['H5'].fill = earth_fill; ws['H5'].alignment = center; ws['H5'].border = thin_border
ws.merge_cells('L5:L6'); ws['L5'] = 'Y2 Total'; ws['L5'].font = white_font; ws['L5'].fill = dark_green_fill; ws['L5'].alignment = center; ws['L5'].border = thin_border

ws.merge_cells('M5:P5'); ws['M5'] = 'Year 3: $271,233'; ws['M5'].font = white_font; ws['M5'].fill = earth_fill; ws['M5'].alignment = center; ws['M5'].border = thin_border
ws.merge_cells('Q5:Q6'); ws['Q5'] = 'Y3 Total'; ws['Q5'].font = white_font; ws['Q5'].fill = dark_green_fill; ws['Q5'].alignment = center; ws['Q5'].border = thin_border

ws.merge_cells('R5:R6'); ws['R5'] = 'Grand Total'; ws['R5'].font = white_font; ws['R5'].fill = dark_green_fill; ws['R5'].alignment = center; ws['R5'].border = thin_border

for col, label in zip(['C','D','E','F','H','I','J','K','M','N','O','P'], ['Q1','Q2','Q3','Q4']*3):
    ws[f'{col}6'] = label
    ws[f'{col}6'].font = white_font_sm; ws[f'{col}6'].fill = dark_green_fill; ws[f'{col}6'].alignment = center; ws[f'{col}6'].border = thin_border

# Activity data: (name, is_sub, y1q[4], y2q[4], y3q[4])
activities = [
    ("System Architecture & Requirements", False, [8000,5000,0,0], [4000,2000,0,0], [3000,0,0,0]),
    ("IoT Sensor Design & GHG Sensor Spec", True, [6000,5000,2000,0], [3000,3000,2000,0], [3000,2000,0,0]),
    ("AI/ML Model Design (Disease, GHG, Yield)", True, [5000,5000,3000,2000], [4000,4000,3000,2000], [3000,3000,3000,2000]),
    ("GHG / Biogas Monitoring System Design", True, [4000,4000,2000,0], [3000,3000,2000,2000], [3000,3000,2000,2000]),
    ("Dashboard & Mobile App UX/UI Design", True, [3000,3000,2000,1000], [2000,2000,2000,1000], [2000,2000,1000,1000]),
    # --- Dev ---
    ("IoT Hardware Procurement & Assembly", False, [3000,10000,8000,4000], [6000,8000,7000,5000], [8000,10000,8000,6000]),
    ("Embedded Software & Firmware Dev", True, [2000,4000,5000,4000], [4000,5000,5000,4000], [5000,6000,6000,5000]),
    ("AI Model Training & Deployment", True, [0,3000,5000,5000], [5000,6000,6000,5000], [7000,7000,7000,6000]),
    ("Web Dashboard & Mobile App Dev", True, [0,3000,5000,5000], [5000,6000,5000,5000], [6000,7000,6000,6000]),
    ("GHG Monitoring, Biogas & NDC Integration", True, [1000,3000,4000,4000], [4000,5000,5000,4000], [5000,6000,6000,5000]),
    ("Cloud Infrastructure, Database & APIs", True, [2000,3000,3000,4479], [4000,4000,4000,4575], [5000,5000,5000,5233]),
    ("Testing, QA & Pilot Deployment", True, [0,0,5000,7000], [4000,5000,5000,5000], [5000,6000,6000,6000]),
    # --- Promo ---
    ("Farmer Training & Capacity Building", False, [0,0,3000,3000], [3000,3000,3000,2000], [3000,3000,2000,2000]),
    ("Stakeholder Engagement & Live Demos", True, [1000,2000,2000,1000], [2000,2000,3000,2000], [2000,2000,2000,2000]),
    ("NDC / GHG Impact Reports & Publications", True, [0,0,2000,3000], [2000,2000,3000,2000], [3000,3000,2000,2000]),
    ("Marketing, Branding & Conferences", True, [2000,2000,3000,2000], [3000,3000,2000,3000], [3000,3000,3000,3000]),
    ("Partnership Dev & Green Fund Applications", True, [2000,3000,3000,4000], [3000,3000,4000,4000], [3000,4000,4000,4000]),
    ("Monitoring & Evaluation (M&E)", True, [1000,0,1000,3000], [1000,2000,2000,3000], [1000,3000,4000,2000]),
]

# Category boundaries: Design=0-4, Dev=5-11, Promo=12-17
cat_info = [
    ("SOLUTION\nDESIGN", 0, 4, "Design Subtotal"),
    ("SOLUTION\nDEVELOPMENT", 5, 11, "Development Subtotal"),
    ("SOLUTION\nPROMOTION", 12, 17, "Promotion Subtotal"),
]

current_row = 7
subtotal_rows = []
cat_row_ranges = []  # (first_data_row, last_data_row) per category

act_idx = 0
for ci, (cat_name, start_i, end_i, sub_name) in enumerate(cat_info):
    first_row = current_row
    for i in range(start_i, end_i + 1):
        name, is_sub, y1q, y2q, y3q = activities[i]
        r = current_row
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2, value=name).font = sub_font if is_sub else header_font
        ws.cell(row=r, column=2).alignment = left_wrap; ws.cell(row=r, column=2).border = thin_border

        # Write numeric values (0 = leave blank)
        for qi, v in enumerate(y1q):
            c = ws.cell(row=r, column=3+qi)
            if v > 0:
                c.value = v
                c.number_format = usd_fmt
            c.font = money_font; c.alignment = right_al; c.border = thin_border

        # Y1 Total formula
        c = ws.cell(row=r, column=7); c.value = f"=SUM(C{r}:F{r})"
        c.font = money_bold; c.alignment = right_al; c.border = thin_border; c.fill = year_total_fill; c.number_format = usd_fmt

        for qi, v in enumerate(y2q):
            c = ws.cell(row=r, column=8+qi)
            if v > 0:
                c.value = v
                c.number_format = usd_fmt
            c.font = money_font; c.alignment = right_al; c.border = thin_border

        # Y2 Total formula
        c = ws.cell(row=r, column=12); c.value = f"=SUM(H{r}:K{r})"
        c.font = money_bold; c.alignment = right_al; c.border = thin_border; c.fill = year_total_fill; c.number_format = usd_fmt

        for qi, v in enumerate(y3q):
            c = ws.cell(row=r, column=13+qi)
            if v > 0:
                c.value = v
                c.number_format = usd_fmt
            c.font = money_font; c.alignment = right_al; c.border = thin_border

        # Y3 Total formula
        c = ws.cell(row=r, column=17); c.value = f"=SUM(M{r}:P{r})"
        c.font = money_bold; c.alignment = right_al; c.border = thin_border; c.fill = year_total_fill; c.number_format = usd_fmt

        # Grand Total formula
        c = ws.cell(row=r, column=18); c.value = f"=G{r}+L{r}+Q{r}"
        c.font = Font(name="Consolas", size=10, bold=True, color="E65100")
        c.alignment = right_al; c.border = thin_border; c.fill = grand_total_fill; c.number_format = usd_fmt

        current_row += 1

    last_row = current_row - 1
    cat_row_ranges.append((first_row, last_row))

    # Subtotal row
    sr = current_row
    subtotal_rows.append(sr)
    ws.cell(row=sr, column=1).border = thin_border; ws.cell(row=sr, column=1).fill = subtotal_fill
    ws.cell(row=sr, column=2, value=sub_name).font = Font(bold=True, size=9, color="2D5016")
    ws.cell(row=sr, column=2).alignment = left_wrap; ws.cell(row=sr, column=2).border = thin_border; ws.cell(row=sr, column=2).fill = subtotal_fill

    for col in range(3, 19):
        cl = get_column_letter(col)
        c = ws.cell(row=sr, column=col)
        c.value = f"=SUM({cl}{first_row}:{cl}{last_row})"
        c.font = Font(name="Consolas", size=9, bold=True, color="2D5016")
        c.alignment = right_al; c.border = thin_border; c.fill = subtotal_fill; c.number_format = usd_fmt

    for col in [7, 12, 17]:
        ws.cell(row=sr, column=col).fill = year_total_fill
    ws.cell(row=sr, column=18).fill = grand_total_fill
    ws.cell(row=sr, column=18).font = Font(name="Consolas", size=10, bold=True, color="E65100")

    # Merge category cell
    ws.merge_cells(f'A{first_row}:A{last_row}')
    ws.cell(row=first_row, column=1).value = cat_name
    ws.cell(row=first_row, column=1).font = Font(bold=True, size=8, color="2D5016")
    ws.cell(row=first_row, column=1).fill = light_green_fill
    ws.cell(row=first_row, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.cell(row=first_row, column=1).border = thin_border

    current_row += 1

# Grand total row
gr = current_row
sr1, sr2, sr3 = subtotal_rows

ws.cell(row=gr, column=1).fill = grand_row_fill; ws.cell(row=gr, column=1).border = thin_border
ws.cell(row=gr, column=2, value="QUARTERLY TOTALS").font = Font(bold=True, size=10, color="FFFFFF")
ws.cell(row=gr, column=2).fill = grand_row_fill
ws.cell(row=gr, column=2).alignment = Alignment(horizontal='right', vertical='center')
ws.cell(row=gr, column=2).border = thin_border

for col in range(3, 19):
    cl = get_column_letter(col)
    c = ws.cell(row=gr, column=col)
    c.value = f"={cl}{sr1}+{cl}{sr2}+{cl}{sr3}"
    c.font = Font(name="Consolas", size=9, bold=True, color="FFFFFF")
    c.fill = grand_row_fill; c.alignment = right_al; c.border = thin_border; c.number_format = usd_fmt

for col in [7, 12, 17]:
    ws.cell(row=gr, column=col).font = Font(name="Consolas", size=10, bold=True, color="F9A825")
ws.cell(row=gr, column=18).font = Font(name="Consolas", size=12, bold=True, color="F9A825")

# Column widths
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 40
for cl in ['C','D','E','F','H','I','J','K','M','N','O','P']:
    ws.column_dimensions[cl].width = 11
for cl in ['G','L','Q']:
    ws.column_dimensions[cl].width = 13
ws.column_dimensions['R'].width = 14

for r in range(5, current_row + 1):
    ws.row_dimensions[r].height = 22
ws.row_dimensions[5].height = 28
ws.row_dimensions[gr].height = 28

ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.orientation = 'landscape'
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.freeze_panes = 'C7'


# ============================================================
# SHEET 2: HARDWARE BUDGET
# ============================================================
ws2 = wb.create_sheet(title="Hardware Budget")

hw_green_fill = PatternFill(start_color="2D5016", end_color="2D5016", fill_type="solid")
hw_light_fill = PatternFill(start_color="F5F9F0", end_color="F5F9F0", fill_type="solid")
hw_total_fill = PatternFill(start_color="E8F0DD", end_color="E8F0DD", fill_type="solid")
hw_footer_fill = PatternFill(start_color="2D5016", end_color="2D5016", fill_type="solid")

# Title
ws2.merge_cells('A1:G1')
ws2['A1'] = "MooMe System: IoT Hardware Cost Estimation (12 Cows)"
ws2['A1'].font = Font(bold=True, size=14, color="2D5016"); ws2['A1'].alignment = center

ws2.merge_cells('A2:G2')
ws2['A2'] = "Prototype Startup Phase | Prices in USD (2,699,000 RWF)"
ws2['A2'].font = Font(size=10, color="7A7A62"); ws2['A2'].alignment = center

# Headers row 4
hw_headers = ['No', 'Component', 'Description', 'Quantity', 'Unit Price (USD)', 'Total (USD)', 'Total (RWF)']
for i, h in enumerate(hw_headers):
    c = ws2.cell(row=4, column=i+1, value=h)
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.fill = hw_green_fill; c.alignment = center; c.border = thin_border

# Hardware data: (no, component, description, qty, unit_rwf)
hardware = [
    (1, "Raspberry Pi 4/5", "Main controller", 1, 120000),
    (2, "MicroSD Card (64GB)", "Storage", 1, 15000),
    (3, "Power Supply", "Adapter", 1, 10000),
    (4, "Non-contact Temp Sensor (IR)", "Per cow", 12, 20000),
    (5, "DHT22 Sensor", "Environment monitoring", 2, 8000),
    (6, "Air Quality Sensor (MQ-135)", "Gas detection", 1, 15000),
    (7, "Oxygen Sensor", "Air monitoring", 1, 25000),
    (8, "RFID Reader Module", "Identification", 1, 18000),
    (9, "RFID Tags", "Per cow ID", 12, 2000),
    (10, "Water Flow Sensor", "Per cow intake", 12, 10000),
    (11, "Load Cell + HX711", "Feed measurement", 12, 15000),
    (12, "Smart Milk Meter", "Milk measurement", 12, 60000),
    (13, "DC Water Pump", "Per cow", 12, 20000),
    (14, "Fans (Cooling)", "Shared system", 4, 25000),
    (15, "Relay Modules", "Control system", 5, 8000),
    (16, "Motor (Feed Dispenser)", "Per cow", 12, 18000),
    (17, "Pipes & Spray Nozzles", "Cooling system", 1, 60000),
    (18, "GSM Module (SIM800L)", "SMS alerts", 1, 15000),
    (19, "Wi-Fi Router", "Connectivity", 1, 35000),
    (20, "Solar Panel (200W)", "Backup power", 1, 150000),
    (21, "Battery (100Ah)", "Energy storage", 1, 180000),
    (22, "Charge Controller", "Solar system", 1, 40000),
    (23, "Inverter", "Power conversion", 1, 80000),
    (24, "Wires & Installation", "Setup", 1, 80000),
]

# Exchange rate cell (so user can change it)
ws2['I4'] = "Exchange Rate:"
ws2['I4'].font = Font(bold=True, size=10, color="2D5016"); ws2['I4'].alignment = Alignment(horizontal='right', vertical='center')
ws2['J4'] = 1174
ws2['J4'].font = Font(bold=True, size=12, color="E65100"); ws2['J4'].number_format = '#,##0'
ws2['K4'] = "RWF / USD"
ws2['K4'].font = Font(size=9, color="7A7A62"); ws2['K4'].alignment = Alignment(horizontal='left', vertical='center')

rate_cell = "$J$4"  # absolute reference

for idx, (no, comp, desc, qty, unit_rwf) in enumerate(hardware):
    r = 5 + idx
    fill = hw_light_fill if idx % 2 == 0 else None

    # No
    c = ws2.cell(row=r, column=1, value=no)
    c.font = Font(size=9); c.alignment = center; c.border = thin_border
    if fill: c.fill = fill

    # Component
    c = ws2.cell(row=r, column=2, value=comp)
    c.font = Font(size=9, bold=True); c.alignment = left_wrap; c.border = thin_border
    if fill: c.fill = fill

    # Description
    c = ws2.cell(row=r, column=3, value=desc)
    c.font = Font(size=9, color="4E4E3A"); c.alignment = left_wrap; c.border = thin_border
    if fill: c.fill = fill

    # Quantity
    c = ws2.cell(row=r, column=4, value=qty)
    c.font = Font(size=9); c.alignment = center; c.border = thin_border
    if fill: c.fill = fill

    # Unit Price USD = unit_rwf / rate (formula)
    c = ws2.cell(row=r, column=5)
    c.value = f"=ROUND(G{r}/D{r}/{rate_cell},0)" if qty > 1 else f"=ROUND(G{r}/{rate_cell},0)"
    c.font = Font(name="Consolas", size=9); c.alignment = right_al; c.border = thin_border; c.number_format = '$#,##0'
    if fill: c.fill = fill

    # Total USD = total_rwf / rate (formula)
    c = ws2.cell(row=r, column=6)
    c.value = f"=ROUND(G{r}/{rate_cell},0)"
    c.font = Font(name="Consolas", size=9, bold=True); c.alignment = right_al; c.border = thin_border; c.number_format = '$#,##0'
    if fill: c.fill = fill

    # Total RWF (source value)
    total_rwf = unit_rwf * qty
    c = ws2.cell(row=r, column=7, value=total_rwf)
    c.font = Font(name="Consolas", size=9, color="7A7A62"); c.alignment = right_al; c.border = thin_border; c.number_format = '#,##0'
    if fill: c.fill = fill

# Total row
total_r = 5 + len(hardware)
first_data = 5
last_data = total_r - 1

ws2.merge_cells(f'A{total_r}:D{total_r}')
c = ws2.cell(row=total_r, column=1, value="Total Prototype Hardware (12 Cows)")
c.font = Font(bold=True, size=11, color="FFFFFF"); c.fill = hw_footer_fill; c.alignment = Alignment(horizontal='right', vertical='center'); c.border = thin_border
for col in [2,3,4]:
    ws2.cell(row=total_r, column=col).fill = hw_footer_fill; ws2.cell(row=total_r, column=col).border = thin_border

# Unit price total (blank/not meaningful)
ws2.cell(row=total_r, column=5).fill = hw_footer_fill; ws2.cell(row=total_r, column=5).border = thin_border

# Total USD = SUM
c = ws2.cell(row=total_r, column=6)
c.value = f"=SUM(F{first_data}:F{last_data})"
c.font = Font(name="Consolas", size=12, bold=True, color="F9A825")
c.fill = hw_footer_fill; c.alignment = right_al; c.border = thin_border; c.number_format = '$#,##0'

# Total RWF = SUM
c = ws2.cell(row=total_r, column=7)
c.value = f"=SUM(G{first_data}:G{last_data})"
c.font = Font(name="Consolas", size=11, bold=True, color="F9A825")
c.fill = hw_footer_fill; c.alignment = right_al; c.border = thin_border; c.number_format = '#,##0'

# Note row
note_r = total_r + 2
ws2.merge_cells(f'A{note_r}:G{note_r}')
ws2.cell(row=note_r, column=1, value="Note: USD prices are calculated using the exchange rate in cell J4. Change J4 to update all USD values automatically.")
ws2.cell(row=note_r, column=1).font = Font(size=9, italic=True, color="7A7A62")
ws2.cell(row=note_r, column=1).alignment = left_wrap

# Column widths
ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 32
ws2.column_dimensions['C'].width = 22
ws2.column_dimensions['D'].width = 10
ws2.column_dimensions['E'].width = 16
ws2.column_dimensions['F'].width = 16
ws2.column_dimensions['G'].width = 16
ws2.column_dimensions['I'].width = 16
ws2.column_dimensions['J'].width = 10
ws2.column_dimensions['K'].width = 12

ws2.row_dimensions[4].height = 28
ws2.row_dimensions[total_r].height = 28

ws2.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws2.page_setup.fitToWidth = 1
ws2.page_setup.fitToHeight = 1
ws2.page_setup.orientation = 'portrait'
ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
ws2.freeze_panes = 'A5'

# ===== SAVE =====
output = "/Users/blaiseai4sense/Desktop/ClaudeRemyProj/MooMe_Logframe_Budget.xlsx"
wb.save(output)
print(f"Excel saved to: {output}")
